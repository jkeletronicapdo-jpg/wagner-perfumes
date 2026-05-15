"""
Wagner Perfumes — Servidor Flask
Landing page promocional + Checkout + Painel Admin
Banco de dados: Excel (.xlsx) — abas: categorias, marcas, produtos
JC Infocell | Agente Desenvolvedor
"""

import json
import os
import io
import base64
import copy
from datetime import datetime
from flask import Flask, render_template, jsonify, request
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, 'dados.xlsx')
PEDIDOS_DIR = os.path.join(BASE_DIR, 'pedidos')
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')

app = Flask(__name__,
            static_folder=os.path.join(os.path.dirname(__file__), '..', 'assets'),
            static_url_path='/static',
            template_folder=os.path.join(os.path.dirname(__file__), '..', 'templates'))

from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


# ==================== LEITURA / ESCRITA XLSX ====================

def _linhas(wb, aba):
    """Retorna lista de dicts de uma aba."""
    ws = wb[aba]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def carregar_categorias():
    wb = openpyxl.load_workbook(XLSX_PATH)
    dados = _linhas(wb, 'categorias')
    wb.close()
    for c in dados:
        if c.get('imagem') is None:
            c['imagem'] = 'cat_masculino.jpg'
    return dados


def carregar_marcas():
    wb = openpyxl.load_workbook(XLSX_PATH)
    dados = _linhas(wb, 'marcas')
    wb.close()
    return dados


def carregar_produtos():
    wb = openpyxl.load_workbook(XLSX_PATH)
    dados = _linhas(wb, 'produtos')
    wb.close()
    for p in dados:
        p['destaque'] = p.get('destaque', 'nao') == 'sim'
    return dados


def salvar_categorias(categorias):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['categorias']
    ws.delete_rows(2, ws.max_row)
    for c in categorias:
        ws.append([c['id'], c['nome']])
    wb.save(XLSX_PATH)
    wb.close()


def salvar_marcas(marcas):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['marcas']
    ws.delete_rows(2, ws.max_row)
    for m in marcas:
        ws.append([m['id'], m['nome']])
    wb.save(XLSX_PATH)
    wb.close()


def salvar_produtos(produtos):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['produtos']
    ws.delete_rows(2, ws.max_row)
    for p in produtos:
        destaque_str = 'sim' if p.get('destaque') else 'nao'
        ws.append([
            p.get('id', 1),
            p.get('nome', ''),
            p.get('marca', ''),
            p.get('categoria', ''),
            p.get('preco', 0),
            p.get('descricao', ''),
            p.get('imagem', ''),
            destaque_str,
            p.get('estoque', 0)
        ])
    wb.save(XLSX_PATH)
    wb.close()


def get_categoria_nome(cat_id, categorias):
    for c in categorias:
        if c['id'] == cat_id:
            return c['nome']
    return cat_id


# ==================== ROTAS PUBLICAS ====================

@app.route('/')
def home():
    categorias = carregar_categorias()
    marcas = carregar_marcas()
    produtos = carregar_produtos()
    for cat in categorias:
        cat['count'] = len([p for p in produtos if p['categoria'] == cat['id']])
    marcas_dict = {m['id']: m['nome'] for m in marcas}
    for p in produtos:
        p['categoria_nome'] = get_categoria_nome(p['categoria'], categorias)
        p['marca_nome'] = marcas_dict.get(p.get('marca', ''), p.get('marca', ''))
    return render_template('index.html', categorias=categorias, produtos=produtos)


@app.route('/checkout')
def checkout():
    return render_template('checkout.html')


@app.route('/admin')
def admin():
    return render_template('admin.html')


# ==================== API CATEGORIAS ====================

@app.route('/api/categorias', methods=['GET'])
def api_listar_categorias():
    return jsonify(carregar_categorias())


@app.route('/api/categoria', methods=['POST'])
def api_criar_categoria():
    try:
        dados = carregar_categorias()
        nova = request.get_json()
        novo_id = nova.get('id', nova['nome'].strip().lower().replace(' ', '_'))
        if any(c['id'] == novo_id for c in dados):
            return jsonify({'status': 'erro', 'mensagem': 'Categoria já existe'}), 400
        dados.append({'id': novo_id, 'nome': nova['nome']})
        salvar_categorias(dados)
        print(f"  [ADMIN] Categoria criada: {nova['nome']} (id={novo_id})")
        return jsonify({'status': 'ok', 'id': novo_id})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/categoria/<cat_id>', methods=['PUT'])
def api_atualizar_categoria(cat_id):
    try:
        dados = carregar_categorias()
        alt = request.get_json()
        for c in dados:
            if c['id'] == cat_id:
                if 'nome' in alt:
                    c['nome'] = alt['nome']
                if 'id' in alt:
                    c['id'] = alt['id']
                salvar_categorias(dados)
                print(f"  [ADMIN] Categoria atualizada: {cat_id}")
                return jsonify({'status': 'ok'})
        return jsonify({'erro': 'Nao encontrada'}), 404
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/categoria/<cat_id>', methods=['DELETE'])
def api_deletar_categoria(cat_id):
    try:
        dados = carregar_categorias()
        dados = [c for c in dados if c['id'] != cat_id]
        salvar_categorias(dados)
        print(f"  [ADMIN] Categoria deletada: {cat_id}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== API MARCAS ====================

@app.route('/api/marcas', methods=['GET'])
def api_listar_marcas():
    return jsonify(carregar_marcas())


@app.route('/api/marca', methods=['POST'])
def api_criar_marca():
    try:
        dados = carregar_marcas()
        nova = request.get_json()
        novo_id = nova.get('id', nova['nome'].strip().lower().replace(' ', '_'))
        if any(m['id'] == novo_id for m in dados):
            return jsonify({'status': 'erro', 'mensagem': 'Marca já existe'}), 400
        dados.append({'id': novo_id, 'nome': nova['nome']})
        salvar_marcas(dados)
        print(f"  [ADMIN] Marca criada: {nova['nome']} (id={novo_id})")
        return jsonify({'status': 'ok', 'id': novo_id})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/marca/<marca_id>', methods=['PUT'])
def api_atualizar_marca(marca_id):
    try:
        dados = carregar_marcas()
        alt = request.get_json()
        for m in dados:
            if m['id'] == marca_id:
                if 'nome' in alt:
                    m['nome'] = alt['nome']
                if 'id' in alt:
                    m['id'] = alt['id']
                salvar_marcas(dados)
                print(f"  [ADMIN] Marca atualizada: {marca_id}")
                return jsonify({'status': 'ok'})
        return jsonify({'erro': 'Nao encontrada'}), 404
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/marca/<marca_id>', methods=['DELETE'])
def api_deletar_marca(marca_id):
    try:
        dados = carregar_marcas()
        dados = [m for m in dados if m['id'] != marca_id]
        salvar_marcas(dados)
        print(f"  [ADMIN] Marca deletada: {marca_id}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== UPLOAD IMAGEM ====================

@app.route('/api/upload-imagem', methods=['POST'])
def api_upload_imagem():
    try:
        data = request.get_json()
        imagem_b64 = data.get('imagem', '')
        nome_arquivo = data.get('nome', '')

        if not imagem_b64 or not nome_arquivo:
            return jsonify({'status': 'erro', 'mensagem': 'Dados incompletos'}), 400

        # Remove prefixo data:image/... se vier
        if ',' in imagem_b64:
            imagem_b64 = imagem_b64.split(',')[1]

        img_data = base64.b64decode(imagem_b64)
        assets_dir = os.path.join(BASE_DIR, '..', 'assets')
        destino = os.path.join(assets_dir, 'img', 'produtos', nome_arquivo)
        os.makedirs(os.path.dirname(destino), exist_ok=True)

        with open(destino, 'wb') as f:
            f.write(img_data)

        print(f"  [UPLOAD] Imagem salva: {nome_arquivo}")
        return jsonify({'status': 'ok', 'arquivo': nome_arquivo, 'url': f'/static/img/produtos/{nome_arquivo}'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== API PRODUTOS ====================

@app.route('/api/produto/<int:produto_id>', methods=['GET'])
def api_get_produto(produto_id):
    produtos = carregar_produtos()
    for p in produtos:
        if p['id'] == produto_id:
            return jsonify(p)
    return jsonify({'erro': 'Produto nao encontrado'}), 404


@app.route('/api/produto', methods=['POST'])
def api_criar_produto():
    try:
        produtos = carregar_produtos()
        novo = request.get_json()
        novo_id = max(p['id'] for p in produtos) + 1 if produtos else 1
        novo['id'] = novo_id
        novo['estoque'] = int(novo.get('estoque', 0))
        novo['destaque'] = novo.get('destaque', False)
        produtos.append(novo)
        salvar_produtos(produtos)
        print(f"  [ADMIN] Produto criado: {novo['nome']} (id={novo_id})")
        return jsonify({'status': 'ok', 'id': novo_id})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/produto/<int:produto_id>', methods=['PUT'])
def api_atualizar_produto(produto_id):
    try:
        produtos = carregar_produtos()
        alt = request.get_json()
        for p in produtos:
            if p['id'] == produto_id:
                for k, v in alt.items():
                    if k != 'id':
                        p[k] = v
                if 'estoque' in alt:
                    p['estoque'] = int(alt['estoque'])
                salvar_produtos(produtos)
                print(f"  [ADMIN] Produto atualizado: {p['nome']} (id={produto_id})")
                return jsonify({'status': 'ok'})
        return jsonify({'erro': 'Nao encontrado'}), 404
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/produto/<int:produto_id>', methods=['DELETE'])
def api_deletar_produto(produto_id):
    try:
        produtos = carregar_produtos()
        produtos = [p for p in produtos if p['id'] != produto_id]
        salvar_produtos(produtos)
        print(f"  [ADMIN] Produto deletado: id={produto_id}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/produtos')
def api_listar_produtos():
    return jsonify(carregar_produtos())


# ==================== API PEDIDOS ====================

@app.route('/api/pedido', methods=['POST'])
def api_criar_pedido():
    try:
        pedido = request.get_json()
        os.makedirs(PEDIDOS_DIR, exist_ok=True)
        pedido_id = f"WP-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        pedido['id'] = pedido_id
        pedido['data'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        pedido['status'] = 'recebido'
        caminho = os.path.join(PEDIDOS_DIR, f'{pedido_id}.json')
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(pedido, f, indent=2, ensure_ascii=False)

        # Atualiza estoque
        produtos = carregar_produtos()
        for item in pedido.get('itens', []):
            for p in produtos:
                if p['id'] == item['id']:
                    p['estoque'] = max(0, int(p.get('estoque', 0)) - item['qty'])
        salvar_produtos(produtos)

        total = sum(item['preco'] * item['qty'] for item in pedido.get('itens', []))
        print(f"\n=== NOVO PEDIDO: {pedido_id} ===")
        print(f"Cliente: {pedido.get('nome')} | Tel: {pedido.get('telefone')}")
        print(f"Total: R$ {total:.2f} | Itens: {len(pedido.get('itens', []))}")
        print("=" * 40)
        return jsonify({'status': 'ok', 'pedido_id': pedido_id, 'total': total})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/pedidos')
def api_listar_pedidos():
    if not os.path.exists(PEDIDOS_DIR):
        return jsonify([])
    pedidos = []
    for fname in sorted(os.listdir(PEDIDOS_DIR), reverse=True):
        if fname.endswith('.json'):
            with open(os.path.join(PEDIDOS_DIR, fname), 'r', encoding='utf-8') as f:
                pedidos.append(json.load(f))
    return jsonify(pedidos)


@app.route('/api/pedido/<pedido_id>/status', methods=['PUT'])
def api_alterar_status(pedido_id):
    try:
        data = request.get_json()
        novo_status = data.get('status', 'recebido')
        caminho = os.path.join(PEDIDOS_DIR, f'{pedido_id}.json')
        if not os.path.exists(caminho):
            return jsonify({'erro': 'Nao encontrado'}), 404
        with open(caminho, 'r', encoding='utf-8') as f:
            pedido = json.load(f)
        pedido['status'] = novo_status
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(pedido, f, indent=2, ensure_ascii=False)
        print(f"  [ADMIN] Pedido {pedido_id} -> {novo_status}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


@app.route('/api/pedido/<pedido_id>', methods=['DELETE'])
def api_deletar_pedido(pedido_id):
    try:
        caminho = os.path.join(PEDIDOS_DIR, f'{pedido_id}.json')
        if os.path.exists(caminho):
            os.remove(caminho)
            print(f"  [ADMIN] Pedido deletado: {pedido_id}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== API PIX ====================

@app.route('/api/pix/gerar', methods=['POST'])
def api_gerar_pix():
    try:
        import qrcode
        dados = request.get_json()
        pedido_id = dados.get('pedido_id', 'WP-000000')
        valor = dados.get('valor', 0)
        payload = f"00020101021226830014BR.GOV.BCB.PIX2563pix.wagnerperfumes.com/{pedido_id}520400005303986540{int(valor):.0f}5802BR5916Wagner Perfumes6008BRASILIA62070503***6304"
        qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=2)
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#3a1f18", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        img_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
        print(f"  [PIX] QR Code gerado para {pedido_id} - R$ {valor:.2f}")
        return jsonify({'status': 'ok', 'qr_code': f"data:image/png;base64,{img_b64}", 'payload': payload, 'copia_e_cola': payload, 'pedido_id': pedido_id, 'valor': valor})
    except ImportError:
        return jsonify({'status': 'erro', 'mensagem': 'QR Code nao disponivel (pip install qrcode[pil])'}), 500
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== API CONFIG ====================

@app.route('/api/config', methods=['GET'])
def api_get_config():
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        cfg = json.load(f)
    return jsonify(cfg)


@app.route('/api/config', methods=['PUT'])
def api_salvar_config():
    try:
        data = request.get_json()
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        cfg.update({k: v for k, v in data.items()})
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, indent=4, ensure_ascii=False)
        print(f"  [ADMIN] Config atualizada")
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500


# ==================== INICIO ====================

if __name__ == '__main__':
    import socket

    print("=" * 50)
    print("  WAGNER PERFUMES — Servidor Web + Admin")
    print("  Banco: Excel (.xlsx)")
    print("=" * 50)

    hostname = socket.gethostname()
    ips = set()
    try:
        for info in socket.getaddrinfo(hostname, None):
            ip = info[4][0]
            if not ip.startswith('127.'):
                ips.add(ip)
    except:
        pass

    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        config = json.load(f)

    cfg = config.get('config', {})
    porta = int(os.environ.get('PORT', cfg.get('porta', 5000)))
    debug = cfg.get('debug', False)

    print(f"\n  Loja:      http://localhost:{porta}")
    print(f"  Admin:     http://localhost:{porta}/admin")
    for ip in sorted(ips):
        print(f"  Rede:      http://{ip}:{porta}")
    print(f"\n  Ctrl+C para parar")
    print("=" * 50)

    app.run(host=cfg.get('host', '0.0.0.0'), port=porta, debug=debug)
