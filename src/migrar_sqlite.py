"""
Migração: Excel (.xlsx) → SQLite
Copia dados do dados.xlsx para database.sqlite
"""
import os, sys
import openpyxl
import sqlite3

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, 'dados.xlsx')
DB_PATH = os.path.join(BASE_DIR, 'database.sqlite')

def migrar():
    if not os.path.exists(XLSX_PATH):
        print("ERRO: dados.xlsx nao encontrado em", XLSX_PATH)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        DROP TABLE IF EXISTS categorias;
        DROP TABLE IF EXISTS marcas;
        DROP TABLE IF EXISTS produtos;
        CREATE TABLE categorias (id TEXT PRIMARY KEY, nome TEXT NOT NULL, imagem TEXT DEFAULT 'cat_masculino.jpg');
        CREATE TABLE marcas (id TEXT PRIMARY KEY, nome TEXT NOT NULL);
        CREATE TABLE produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            marca TEXT,
            categoria TEXT,
            preco REAL NOT NULL DEFAULT 0,
            descricao TEXT DEFAULT '',
            imagem TEXT DEFAULT 'perfume_png_01.png',
            destaque INTEGER DEFAULT 0,
            estoque INTEGER DEFAULT 0
        );
    """)

    # Categorias
    ws = wb['categorias']
    headers = [c.value for c in ws[1]]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        dados = {headers[i]: row[i] for i in range(len(headers))}
        conn.execute("INSERT INTO categorias (id, nome, imagem) VALUES (?, ?, ?)",
                     (dados['id'], dados['nome'], dados.get('imagem', 'cat_masculino.jpg')))
        count += 1
    print(f"  Categorias: {count} importadas")

    # Marcas
    ws = wb['marcas']
    headers = [c.value for c in ws[1]]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        dados = {headers[i]: row[i] for i in range(len(headers))}
        conn.execute("INSERT INTO marcas (id, nome) VALUES (?, ?)", (dados['id'], dados['nome']))
        count += 1
    print(f"  Marcas: {count} importadas")

    # Produtos
    ws = wb['produtos']
    headers = [c.value for c in ws[1]]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        dados = {headers[i]: row[i] for i in range(len(headers))}
        destaque = 1 if str(dados.get('destaque', 'nao')).lower() in ('sim', 'true', '1') else 0
        conn.execute(
            "INSERT INTO produtos (id, nome, marca, categoria, preco, descricao, imagem, destaque, estoque) VALUES (?,?,?,?,?,?,?,?,?)",
            (dados['id'], dados['nome'], dados.get('marca', ''), dados.get('categoria', ''),
             float(dados.get('preco', 0)), dados.get('descricao', ''), dados.get('imagem', 'perfume_png_01.png'),
             destaque, int(dados.get('estoque', 0))))
        count += 1
    print(f"  Produtos: {count} importados")

    conn.commit()
    conn.close()
    wb.close()
    print(f"\nMigracao concluida! Banco salvo em: {DB_PATH}")

if __name__ == '__main__':
    migrar()
